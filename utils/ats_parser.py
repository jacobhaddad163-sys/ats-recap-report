"""
ATS (Available to Ship) report parser.

Parses raw ATS Excel files pasted from the internal system.
Auto-detects column layout from the STYLE header row so it works
with any ATS report format (Nike, Jordan, Hurley, etc.).

Two format detection:
  Format A: Sheets with TODDLER/4-7 summary headers (one column before OH).
    - Row with "TODDLER" has summary OH/WIP/TOTAL
    - Row below has category name in A, "4-7", and summary OH/WIP/TOTAL
  Format B: Sheets without TODDLER/4-7 summary headers.
    - Category names from non-empty cells in column A
    - OH/WIP computed by summing individual style rows grouped by size prefix

REF# extraction from style codes:
  - Strip first 2 characters (size prefix): 76F610-C5E-P3 -> F610-C5E-P3
  - Take everything before the first dash: F610-C5E-P3 -> F610

Size range from first digit of style code (all 10 ranges):
  0=NB GIRL, 1=INFANT GIRL, 2=TODDLER GIRL, 3=4-6X GIRL, 4=7-16 GIRL
  5=NB BOY, 6=INFANT BOY, 7=TODDLER BOY, 8=4-7 BOY, 9=8-20 BOY
"""

import io
import logging
import re
import xml.etree.ElementTree as _ET
import zipfile
from collections import defaultdict, OrderedDict
from typing import Dict, List

import openpyxl
try:
    from PIL import Image
except ImportError:
    Image = None

logger = logging.getLogger(__name__)
MAX_IMAGE_SIZE = 10 * 1024 * 1024

BRAND_KEYWORDS = {
    "JORDAN": "JORDAN", "NIKE": "NIKE", "HURLEY": "HURLEY",
    "CONVERSE": "CONVERSE", "LEVIS": "LEVIS",
    "UMBRO": "UMBRO", "UNDER ARMOUR": "UNDER ARMOUR",
    "CHAMPION": "CHAMPION", "REEBOK": "REEBOK",
}

# Sheet name -> brand mapping for known sheet names
SHEET_BRAND_MAP = {
    "LONG BOTTOMS": "NIKE LONG BOTTOMS",
    "BOTTOMS": "NIKE LONG BOTTOMS",
    "NIKE TEES": "NIKE TEES",
    "JORDAN TEES": "JORDAN TEES",
}

# Desired processing order for known brand labels
BRAND_ORDER = ["NIKE LONG BOTTOMS", "NIKE TEES", "JORDAN TEES"]


def detect_brand(text: str) -> str:
    t = text.upper()
    for kw, brand in BRAND_KEYWORDS.items():
        if kw in t:
            return brand
    return ""


def map_sheet_to_brand(sheet_name: str) -> str:
    """Map a sheet name to a specific brand label using known mappings.

    Returns the mapped brand label if found, otherwise returns empty string
    (caller should fall back to detect_brand).
    """
    sn = sheet_name.upper().strip()
    # Check exact matches first
    if sn in SHEET_BRAND_MAP:
        return SHEET_BRAND_MAP[sn]
    # Check if sheet name contains known brand+category combos
    for key, brand_label in SHEET_BRAND_MAP.items():
        if key in sn:
            return brand_label
    return ""


def ref_from_style(style_raw: str) -> str:
    """
    Extract REF# from style code.
    1. Strip first 2 chars (size prefix): 76F610-C5E-P3 -> F610-C5E-P3
    2. Take everything before first dash: F610-C5E-P3 -> F610
    """
    style_raw = style_raw.strip()
    if len(style_raw) < 3:
        return style_raw
    stripped = style_raw[2:]  # Remove first 2 chars
    ref = stripped.split("-")[0]  # Take before first dash
    return ref


SIZE_RANGE_MAP = {
    '0': "NB GIRL",
    '1': "INFANT GIRL",
    '2': "TODDLER GIRL",
    '3': "4-6X GIRL",
    '4': "7-16 GIRL",
    '5': "NB BOY",
    '6': "INFANT BOY",
    '7': "TODDLER BOY",
    '8': "4-7 BOY",
    '9': "8-20 BOY",
}


def size_range_from_style(style_num: str) -> str:
    """
    First digit of style number = size range code:
      0=NB Girl, 1=Inf Girl, 2=Tod Girl, 3=4-6X Girl, 4=7-16 Girl
      5=NB Boy, 6=Inf Boy, 7=Toddler Boy, 8=4-7 Boy, 9=8-20 Boy
    """
    if not style_num:
        return "UNKNOWN"
    first = style_num[0]
    return SIZE_RANGE_MAP.get(first, "UNKNOWN")


def _safe_num(v, default=0) -> int:
    if v is None:
        return default
    s = str(v).strip()
    if not s or s.startswith("#"):
        return default
    if re.match(r"\d{1,2}/\d{1,2}/\d{2,4}", s):
        return default
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return default


def _safe_float(v, default=0.0) -> float:
    if v is None:
        return default
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return default


def _safe_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.startswith("#") else s


def _detect_columns(ws) -> dict:
    """Auto-detect column layout by finding the STYLE header row.

    Scans for the first row containing a cell with value "STYLE" (case-insensitive).
    From that row, maps: STYLE, COLOR, ON HAND, WIP, AVAILABILITY, MSRP columns.
    Size columns = everything between COLOR and ON HAND.

    Returns dict with keys:
        style, color, oh, wip, avail, msrp, size_start, size_end, header_row,
        summary_label_col (one column before OH, where TODDLER/4-7 labels go)
    Falls back to Nike/Jordan defaults if no header row found.
    """
    # Default layout (Nike/Jordan): STYLE=3, COLOR=4, sizes=5-11, OH=12, WIP=13, AVAIL=14, MSRP=15
    defaults = {
        "style": 3, "color": 4, "oh": 12, "wip": 13, "avail": 14, "msrp": 15,
        "size_start": 5, "size_end": 11, "header_row": 0,
        "summary_label_col": 11,
    }

    for row_num in range(1, min(ws.max_row + 1, 200)):  # Don't scan forever
        for col_num in range(1, min(ws.max_column + 1, 40)):
            val = ws.cell(row=row_num, column=col_num).value
            if val is not None and str(val).strip().upper() in ("STYLE", "STYLE #", "STYLE NUMBER"):
                # Found the header row. Now map all columns from this row.
                header_map = {}
                for c in range(1, min(ws.max_column + 1, 40)):
                    cv = ws.cell(row=row_num, column=c).value
                    if cv is not None:
                        header_map[str(cv).strip().upper()] = c

                style_col = col_num
                color_col = header_map.get("COLOR", header_map.get("COLOR NAME", style_col + 1))

                # ON HAND can appear as "ON HAND" or "OH"
                oh_col = header_map.get("ON HAND", header_map.get("OH", 0))
                wip_col = header_map.get("WIP", 0)
                avail_col = header_map.get("AVAILABILITY", header_map.get("AVAIL", 0))
                msrp_col = header_map.get("MSRP", 0)

                if not oh_col:
                    # Couldn't find OH column, fall back to defaults
                    logger.warning(f"Header row {row_num} found but no ON HAND/OH column. Using defaults.")
                    return defaults

                # Size columns: everything between COLOR and ON HAND
                size_start = color_col + 1
                size_end = oh_col - 1

                # Summary label column: one before OH (where TODDLER/4-7 labels appear)
                summary_label_col = oh_col - 1

                result = {
                    "style": style_col,
                    "color": color_col,
                    "oh": oh_col,
                    "wip": wip_col if wip_col else oh_col + 1,
                    "avail": avail_col if avail_col else (wip_col + 1 if wip_col else oh_col + 2),
                    "msrp": msrp_col if msrp_col else (avail_col + 1 if avail_col else oh_col + 3),
                    "size_start": size_start,
                    "size_end": size_end,
                    "header_row": row_num,
                    "summary_label_col": summary_label_col,
                }
                logger.info(f"Detected columns from header row {row_num}: {result}")
                return result

    logger.warning("No STYLE header row found. Using default Nike/Jordan column layout.")
    return defaults


def _is_skip_row(style_val: str) -> bool:
    """Check if a row should be skipped for REF# extraction and style parsing.

    Skip rows where the STYLE column value is:
    - "STYLE" (or variants)
    - "TOTAL :" (or starts with TOTAL)
    - blank
    - color legend (contains " - " like "023 - BLACK")
    Also skips common extra header labels like "YEAR", "SEASON".
    """
    if not style_val:
        return True
    upper = style_val.upper().strip()
    if upper in ("STYLE", "STYLE #", "STYLE NUMBER", "YEAR", "SEASON",
                  "SIZE SCALE", "COLOR", "COLOR NAME", "ON HAND", "OH",
                  "WIP", "AVAILABILITY", "MSRP"):
        return True
    if upper.startswith("TOTAL"):
        return True
    if " - " in style_val:
        return True
    return False


# --- Image extraction (unchanged) -----------------------------------------------

def _is_yellow(cell) -> bool:
    """Check if a cell has yellow fill (category header)."""
    try:
        if cell.fill and cell.fill.start_color:
            rgb = str(cell.fill.start_color.rgb or '').upper()
            return 'FFFF00' in rgb
    except Exception:
        pass
    return False


def _safe_zip_path(target: str, prefix: str = 'xl/') -> str:
    cleaned = target
    while '../' in cleaned:
        cleaned = cleaned.replace('../', '')
    cleaned = cleaned.lstrip('/')
    if not cleaned.startswith(prefix):
        cleaned = prefix + cleaned
    if '..' in cleaned:
        return ''
    return cleaned


def _extract_images(file_bytes: bytes) -> Dict[str, Dict[int, bytes]]:
    result = {}
    _NS_REL = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
    _NS_XDR = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    }
    _NS_WB = {
        'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    }
    _R_EMBED = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            names = set(zf.namelist())
            wb_root = _ET.fromstring(zf.read('xl/workbook.xml'))
            sheet_rid = {}
            for sh in wb_root.findall('.//x:sheet', _NS_WB):
                rid, name = sh.get(_R_EMBED, ''), sh.get('name', '')
                if rid and name:
                    sheet_rid[rid] = name

            if 'xl/_rels/workbook.xml.rels' not in names:
                return result
            rels_root = _ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            rid_to_file = {r.get('Id', ''): r.get('Target', '') for r in rels_root.findall('r:Relationship', _NS_REL)}

            for rid, sheet_name in sheet_rid.items():
                ws_file = rid_to_file.get(rid, '')
                if not ws_file:
                    continue
                ws_rels = f'xl/worksheets/_rels/{ws_file.replace("worksheets/", "")}.rels'
                if ws_rels not in names:
                    continue
                ws_rels_root = _ET.fromstring(zf.read(ws_rels))
                drawing_file = None
                for rel in ws_rels_root.findall('r:Relationship', _NS_REL):
                    if 'drawing' in rel.get('Type', '').lower():
                        drawing_file = _safe_zip_path(rel.get('Target', ''), 'xl/')
                        break
                if not drawing_file or drawing_file not in names:
                    continue
                draw_rels = drawing_file.replace('/drawings/', '/drawings/_rels/') + '.rels'
                img_rid_map = {}
                if draw_rels in names:
                    for rel in _ET.fromstring(zf.read(draw_rels)).findall('r:Relationship', _NS_REL):
                        if 'image' in rel.get('Type', '').lower():
                            p = _safe_zip_path(rel.get('Target', ''), 'xl/')
                            if p:
                                img_rid_map[rel.get('Id', '')] = p
                if not img_rid_map:
                    continue
                draw_root = _ET.fromstring(zf.read(drawing_file))
                product_imgs = {}
                for tag in ('xdr:oneCellAnchor', 'xdr:twoCellAnchor'):
                    for anchor in draw_root.findall(tag, _NS_XDR):
                        fr = anchor.find('xdr:from', _NS_XDR)
                        if fr is None:
                            continue
                        re_ = fr.find('xdr:row', _NS_XDR)
                        if re_ is None:
                            continue
                        row_1 = int(re_.text or '0') + 1
                        blip = anchor.find('.//a:blip', _NS_XDR)
                        if blip is None:
                            continue
                        irid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed', '')
                        ipath = img_rid_map.get(irid, '')
                        if not ipath or ipath not in names:
                            continue
                        info = zf.getinfo(ipath)
                        if info.file_size > MAX_IMAGE_SIZE:
                            continue
                        data = zf.read(ipath)
                        try:
                            if Image is None:
                                product_imgs[row_1] = data
                                continue
                            img = Image.open(io.BytesIO(data))
                            if img.size[0] > 100 and img.size[1] > 100:
                                product_imgs[row_1] = data
                        except Exception:
                            pass
                if product_imgs:
                    result[sheet_name] = product_imgs
    except Exception:
        pass
    return result


# --- Format detection helpers ---------------------------------------------------

def _detect_format_a(ws, cols: dict) -> bool:
    """Detect Format A: scan summary_label_col for cells containing exactly 'TODDLER'."""
    label_col = cols["summary_label_col"]
    for row_num in range(1, ws.max_row + 1):
        val = ws.cell(row=row_num, column=label_col).value
        if val is not None and str(val).strip().upper() == "TODDLER":
            return True
    return False


def _extract_refs_between_rows(ws, start_row: int, end_row: int,
                               cols: dict) -> Dict[str, List[str]]:
    """Extract unique REF#s from style codes between start_row and end_row.

    Uses detected column layout (cols dict) for style, OH, WIP, and size columns.
    Returns dict {size_range_name: [refs]} in order encountered.
    Skips STYLE, TOTAL, blank, and color legend rows.
    Also skips ratio rows (rows without OH value).
    """
    style_col = cols["style"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    size_start = cols["size_start"]
    size_end = cols["size_end"]

    refs_by_sr = {}  # {size_range: [refs]}

    for r in range(start_row, end_row + 1):
        c_val = _safe_str(ws.cell(row=r, column=style_col).value)
        if _is_skip_row(c_val):
            continue
        # Must have digits to be a style code
        if not any(ch.isdigit() for ch in c_val):
            continue
        # Skip ratio rows: rows without an OH value
        oh_val = ws.cell(row=r, column=oh_col).value
        if oh_val is None:
            continue
        oh = _safe_num(oh_val)
        wip = _safe_num(ws.cell(row=r, column=wip_col).value)
        # Only count if it looks like a data row (has OH or WIP)
        if oh == 0 and wip == 0:
            # Check if any size cells have data (might still be a valid row with 0s)
            has_size_data = False
            for ci in range(size_start, size_end + 1):
                cv = ws.cell(row=r, column=ci).value
                if cv is not None and str(cv).strip():
                    has_size_data = True
                    break
            if not has_size_data:
                continue

        ref = ref_from_style(c_val)
        sr = size_range_from_style(c_val)
        if ref and sr != "UNKNOWN":
            if sr not in refs_by_sr:
                refs_by_sr[sr] = []
            if ref not in refs_by_sr[sr]:
                refs_by_sr[sr].append(ref)

    return refs_by_sr


def _parse_format_a(ws, ws_fmt, images_by_row: dict, cols: dict) -> List[dict]:
    """Parse Format A sheet: has TODDLER/4-7 summary headers.

    Category detection: When 'TODDLER' found at row N in summary_label_col:
      - Row N = TODDLER summary: OH/WIP/Total in detected columns
      - Row N+1 = BOYS 4-7 summary: col A = CATEGORY NAME, label='4-7', OH/WIP/Total
    Only include categories with the standard TODDLER/4-7 header pattern.

    Format A maps "TODDLER" -> "TODDLER BOY" and "4-7" -> "4-7 BOY" in the
    new size_ranges structure.
    """
    label_col = cols["summary_label_col"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    avail_col = cols["avail"]

    # Find all TODDLER rows in summary_label_col
    toddler_rows = []
    for row_num in range(1, ws.max_row + 1):
        val = ws.cell(row=row_num, column=label_col).value
        if val is not None and str(val).strip().upper() == "TODDLER":
            toddler_rows.append(row_num)

    if not toddler_rows:
        return []

    # Build list of validated category summaries (toddler_row, b47_row, name, OH/WIP)
    cat_summaries = []
    for tod_row in toddler_rows:
        b47_row = tod_row + 1
        cat_name_raw = _safe_str(ws.cell(row=b47_row, column=1).value)
        k_val = _safe_str(ws.cell(row=b47_row, column=label_col).value)

        if "4-7" not in k_val.upper():
            logger.warning(f"Format A: Row {b47_row} col {label_col}='{k_val}', expected '4-7'. Skipping.")
            continue

        tod_oh = _safe_num(ws.cell(row=tod_row, column=oh_col).value)
        tod_wip = _safe_num(ws.cell(row=tod_row, column=wip_col).value)
        tod_total = _safe_num(ws.cell(row=tod_row, column=avail_col).value)
        if tod_total == 0 and (tod_oh > 0 or tod_wip > 0):
            tod_total = tod_oh + tod_wip

        b47_oh = _safe_num(ws.cell(row=b47_row, column=oh_col).value)
        b47_wip = _safe_num(ws.cell(row=b47_row, column=wip_col).value)
        b47_total = _safe_num(ws.cell(row=b47_row, column=avail_col).value)
        if b47_total == 0 and (b47_oh > 0 or b47_wip > 0):
            b47_total = b47_oh + b47_wip

        cat_summaries.append({
            "tod_row": tod_row, "b47_row": b47_row,
            "name": cat_name_raw.rstrip(),
            "tod_oh": tod_oh, "tod_wip": tod_wip, "tod_total": tod_total,
            "b47_oh": b47_oh, "b47_wip": b47_wip, "b47_total": b47_total,
        })

    categories = []

    for i, cs in enumerate(cat_summaries):
        # Data lives AFTER the 4-7 summary row, up to the next TODDLER row (or end of sheet)
        data_start = cs["b47_row"] + 1
        if i + 1 < len(cat_summaries):
            # Stop before next category's sub-header row (OH/WIP/TOTAL label, 2 rows before TODDLER)
            data_end = cat_summaries[i + 1]["tod_row"] - 1
        else:
            data_end = ws.max_row

        # Extract REF#s from style rows in this range
        refs_by_sr = _extract_refs_between_rows(ws, data_start, data_end, cols)

        # Parse blocks in this range
        blocks = _parse_blocks_in_range(ws, data_start, data_end, images_by_row, cols)

        tod_oh, tod_wip, tod_total = cs["tod_oh"], cs["tod_wip"], cs["tod_total"]
        b47_oh, b47_wip, b47_total = cs["b47_oh"], cs["b47_wip"], cs["b47_total"]

        # Build size_ranges dict — Format A maps TODDLER -> TODDLER BOY, 4-7 -> 4-7 BOY
        size_ranges = OrderedDict()
        if tod_oh > 0 or tod_wip > 0 or tod_total > 0 or refs_by_sr.get("TODDLER BOY"):
            size_ranges["TODDLER BOY"] = {
                "oh": tod_oh, "wip": tod_wip, "total": tod_total,
                "refs": refs_by_sr.get("TODDLER BOY", []),
            }
        if b47_oh > 0 or b47_wip > 0 or b47_total > 0 or refs_by_sr.get("4-7 BOY"):
            size_ranges["4-7 BOY"] = {
                "oh": b47_oh, "wip": b47_wip, "total": b47_total,
                "refs": refs_by_sr.get("4-7 BOY", []),
            }

        categories.append({
            "name": cs["name"],
            "size_ranges": size_ranges,
            "blocks": blocks,
        })

    return categories


def _parse_format_b(ws, ws_fmt, images_by_row: dict, cols: dict) -> List[dict]:
    """Parse Format B sheet: no TODDLER/4-7 summary headers.

    Category detection: Scan column A for non-empty cells that are NOT
    'ATS RECAP' and NOT blank. Each such cell is a category name.

    OH/WIP: SUM individual style detail rows grouped by size range (all 10).
    """
    style_col = cols["style"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]

    # Find category rows in column A
    category_rows = []  # [(row_num, category_name)]
    for row_num in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=1).value
        if cell_val is None:
            continue
        val_str = str(cell_val).strip()
        if not val_str:
            continue
        upper = val_str.upper()
        if upper in ("ATS RECAP", "ATS", "RECAP", "OH", "WIP", "TOTAL"):
            continue
        # Use yellow cell detection if formatting workbook available
        cell_fmt = ws_fmt.cell(row=row_num, column=1) if ws_fmt else None
        if cell_fmt and _is_yellow(cell_fmt):
            category_rows.append((row_num, val_str.rstrip()))
        elif not cell_fmt:
            # No formatting available, treat any non-empty col A as category
            category_rows.append((row_num, val_str.rstrip()))

    if not category_rows:
        return []

    categories = []

    for cat_idx, (cat_row, cat_name) in enumerate(category_rows):
        data_start = cat_row + 1
        data_end = category_rows[cat_idx + 1][0] - 1 if cat_idx + 1 < len(category_rows) else ws.max_row

        # Extract REF#s
        refs_by_sr = _extract_refs_between_rows(ws, data_start, data_end, cols)

        # Parse blocks
        blocks = _parse_blocks_in_range(ws, data_start, data_end, images_by_row, cols)

        # Compute OH/WIP from individual style rows by size range
        sr_oh = defaultdict(int)
        sr_wip = defaultdict(int)

        for r in range(data_start, data_end + 1):
            c_val = _safe_str(ws.cell(row=r, column=style_col).value)
            if _is_skip_row(c_val):
                continue
            if not any(ch.isdigit() for ch in c_val):
                continue
            # Skip ratio rows (no OH value)
            oh_raw = ws.cell(row=r, column=oh_col).value
            if oh_raw is None:
                continue

            oh = _safe_num(oh_raw)
            wip = _safe_num(ws.cell(row=r, column=wip_col).value)

            sr = size_range_from_style(c_val)
            if sr != "UNKNOWN":
                sr_oh[sr] += oh
                sr_wip[sr] += wip

        # Build size_ranges dict — only include size ranges with data or refs
        size_ranges = OrderedDict()
        # Collect all size range names from both OH/WIP data and refs
        all_sr_names = list(OrderedDict.fromkeys(
            list(sr_oh.keys()) + list(sr_wip.keys()) + list(refs_by_sr.keys())
        ))
        for sr_name in all_sr_names:
            oh_val = sr_oh.get(sr_name, 0)
            wip_val = sr_wip.get(sr_name, 0)
            refs = refs_by_sr.get(sr_name, [])
            if oh_val > 0 or wip_val > 0 or refs:
                size_ranges[sr_name] = {
                    "oh": oh_val, "wip": wip_val, "total": oh_val + wip_val,
                    "refs": refs,
                }

        categories.append({
            "name": cat_name,
            "size_ranges": size_ranges,
            "blocks": blocks,
        })

    return categories


def _merge_same_name_categories(categories: List[dict]) -> List[dict]:
    """Merge categories with the same name into one entry, preserving order."""
    merged = OrderedDict()
    for cat in categories:
        name = cat["name"]
        if name in merged:
            existing = merged[name]
            # Merge size_ranges
            for sr_name, sr_data in cat["size_ranges"].items():
                if sr_name in existing["size_ranges"]:
                    ex_sr = existing["size_ranges"][sr_name]
                    ex_sr["oh"] += sr_data["oh"]
                    ex_sr["wip"] += sr_data["wip"]
                    ex_sr["total"] += sr_data["total"]
                    for ref in sr_data["refs"]:
                        if ref not in ex_sr["refs"]:
                            ex_sr["refs"].append(ref)
                else:
                    existing["size_ranges"][sr_name] = {
                        "oh": sr_data["oh"], "wip": sr_data["wip"],
                        "total": sr_data["total"], "refs": list(sr_data["refs"]),
                    }
            existing["blocks"].extend(cat["blocks"])
        else:
            # Deep copy size_ranges to avoid mutation
            sr_copy = OrderedDict()
            for sr_name, sr_data in cat["size_ranges"].items():
                sr_copy[sr_name] = {
                    "oh": sr_data["oh"], "wip": sr_data["wip"],
                    "total": sr_data["total"], "refs": list(sr_data["refs"]),
                }
            merged[name] = {**cat, "size_ranges": sr_copy}
    return list(merged.values())


def _parse_blocks_in_range(ws, start_row: int, end_row: int,
                           images_by_row: dict, cols: dict) -> list:
    """Parse style blocks within a row range. Shared by both Format A and Format B.

    Uses detected column layout (cols dict) for all column positions.
    """
    style_col = cols["style"]
    color_col = cols["color"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    avail_col = cols["avail"]
    msrp_col = cols["msrp"]
    size_start = cols["size_start"]
    size_end = cols["size_end"]

    blocks = []
    row = start_row

    def _is_total_row(r):
        """Check if row r is a TOTAL row by checking the style col and col 3."""
        for check_col in set([style_col, 3]):
            v = _safe_str(ws.cell(row=r, column=check_col).value)
            if v.upper().startswith("TOTAL"):
                return True
        return False

    def _is_header_row(r):
        """Check if row r is a STYLE header row."""
        v = _safe_str(ws.cell(row=r, column=style_col).value)
        return v.upper() in ("STYLE", "STYLE #", "STYLE NUMBER")

    while row <= end_row:
        # Sub-header row (STYLE, COLOR, SIZE SCALE)
        if _is_header_row(row):
            block_header_row = row
            row += 1
            block_rows = []

            while row <= end_row:
                if _is_total_row(row):
                    total_oh = _safe_num(ws.cell(row=row, column=oh_col).value)
                    total_wip = _safe_num(ws.cell(row=row, column=wip_col).value)

                    block_ref = block_rows[0]["ref_num"] if block_rows else ""

                    # Product image
                    product_img = None
                    for lr in range(block_header_row, max(0, block_header_row - 15), -1):
                        if lr in images_by_row:
                            product_img = images_by_row[lr]
                            break

                    blocks.append({
                        "ref_num": block_ref,
                        "rows": block_rows,
                        "total_oh": total_oh,
                        "total_wip": total_wip,
                        "product_image": product_img,
                    })

                    row += 1
                    break

                # Style data row — check the style column for a style code
                cv = _safe_str(ws.cell(row=row, column=style_col).value)

                if cv and any(ch.isdigit() for ch in cv) and not _is_skip_row(cv):
                    color = _safe_str(ws.cell(row=row, column=color_col).value)
                    oh = _safe_num(ws.cell(row=row, column=oh_col).value)
                    wip = _safe_num(ws.cell(row=row, column=wip_col).value)
                    avail = _safe_str(ws.cell(row=row, column=avail_col).value)
                    msrp = _safe_float(ws.cell(row=row, column=msrp_col).value)

                    is_label_row = oh > 0 or wip > 0
                    for ci in range(size_start, size_end + 1):
                        cval = ws.cell(row=row, column=ci).value
                        if cval and isinstance(cval, str) and 'T' in cval.upper():
                            is_label_row = True
                            break

                    ref = ref_from_style(cv)
                    sr = size_range_from_style(cv)

                    cells = {}
                    for ci in range(size_start, size_end + 1):
                        cval = ws.cell(row=row, column=ci).value
                        if cval is not None:
                            cells[ci] = cval

                    block_rows.append({
                        "style_num": cv, "ref_num": ref, "color": color,
                        "cells": cells, "oh": oh, "wip": wip,
                        "availability": avail, "msrp": msrp,
                        "size_range": sr, "is_label_row": is_label_row,
                    })
                row += 1
        else:
            row += 1

    return blocks


# --- Sheet ordering helper -----------------------------------------------------

def _sort_sheets_for_processing(sheet_names: List[str]) -> List[str]:
    """Sort sheets so known brands come in the desired order.

    Order: NIKE LONG BOTTOMS first, then NIKE TEES, then JORDAN TEES,
    then any remaining sheets in original order.
    """
    ordered = []
    remaining = list(sheet_names)

    for brand_label in BRAND_ORDER:
        for sn in list(remaining):
            mapped = map_sheet_to_brand(sn)
            if mapped == brand_label:
                ordered.append(sn)
                remaining.remove(sn)

    # Append any sheets that didn't match a known brand order
    ordered.extend(remaining)
    return ordered


# --- Main Parser ---------------------------------------------------------------

def parse_ats_file(file_bytes: bytes) -> dict:
    """
    Parse raw ATS Excel file.

    Returns:
    {
        "sheets": [
            {
                "name": str,
                "brand": str,
                "categories": list of {
                    "name": str,
                    "size_ranges": OrderedDict({
                        "TODDLER BOY": {"oh": int, "wip": int, "total": int, "refs": [str]},
                        "4-7 BOY": {"oh": int, "wip": int, "total": int, "refs": [str]},
                        ...
                    }),
                    "blocks": [block_dict, ...],
                },
                "all_ref_nums": [str, ...],
            },
            ...
        ]
    }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    # Second load for formatting (yellow cell detection)
    wb_fmt = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    all_images = _extract_images(file_bytes)

    result = {"sheets": [], "logo_image": None}

    # Extract logo: the Haddad Brands logo is typically at row 1 of the first sheet
    for sheet_name_check in wb.sheetnames:
        logo_imgs = all_images.get(sheet_name_check, {})
        if 1 in logo_imgs:
            logo_data = logo_imgs[1]
            try:
                if Image is not None:
                    logo_img = Image.open(io.BytesIO(logo_data))
                    # Logo is wide (width > height) and small, not a product image
                    if logo_img.size[0] > logo_img.size[1] and logo_img.size[1] < 200:
                        result["logo_image"] = logo_data
                        break
            except Exception:
                pass

    # Step 1: Identify raw data sheets - exclude "RECAP SHEET" (case-insensitive)
    data_sheet_names = []
    for sheet_name in wb.sheetnames:
        sn_upper = sheet_name.upper().strip()
        if 'RECAP' in sn_upper and 'SHEET' in sn_upper:
            continue
        if sn_upper in ('RECAP', 'SUMMARY', 'RECAP SHEET'):
            continue
        data_sheet_names.append(sheet_name)

    # Sort sheets for processing order
    data_sheet_names = _sort_sheets_for_processing(data_sheet_names)

    for sheet_name in data_sheet_names:
        ws = wb[sheet_name]
        ws_fmt = wb_fmt[sheet_name]

        # Brand detection: try specific sheet-to-brand mapping first, fall back to generic
        brand = map_sheet_to_brand(sheet_name)
        if not brand:
            brand = detect_brand(sheet_name)

        images_by_row = all_images.get(sheet_name, {})

        # Auto-detect column layout from the STYLE header row
        cols = _detect_columns(ws)

        # Step 2: Detect format and parse accordingly
        is_format_a = _detect_format_a(ws, cols)

        if is_format_a:
            logger.info(f"Sheet '{sheet_name}': Format A detected (TODDLER/4-7 headers)")
            categories = _parse_format_a(ws, ws_fmt, images_by_row, cols)
        else:
            logger.info(f"Sheet '{sheet_name}': Format B detected (no summary headers)")
            categories = _parse_format_b(ws, ws_fmt, images_by_row, cols)

        # Collect all ref nums from all size ranges
        all_ref_nums = []
        for cat in categories:
            for sr_data in cat["size_ranges"].values():
                for r in sr_data["refs"]:
                    if r not in all_ref_nums:
                        all_ref_nums.append(r)

        if categories:
            logger.info(f"Sheet '{sheet_name}': {len(categories)} categories: "
                        f"{[c['name'] for c in categories]}")

        result["sheets"].append({
            "name": sheet_name, "brand": brand,
            "categories": categories, "all_ref_nums": all_ref_nums,
            "columns": cols,
        })

    return result


# --- Filter --------------------------------------------------------------------

def filter_blocks(blocks: list, min_units: int = 120, max_units: int = None) -> list:
    """Filter style packs by OH + WIP threshold. Returns filtered blocks."""
    filtered = []
    for block in blocks:
        frows = []
        i = 0
        rows = block["rows"]
        while i < len(rows):
            r = rows[i]
            if r.get("is_label_row", True):
                total = r["oh"] + r["wip"]
                ratio = None
                if i + 1 < len(rows):
                    nxt = rows[i + 1]
                    if nxt.get("style_num") == r.get("style_num") and not nxt.get("is_label_row", True):
                        ratio = nxt
                keep = total >= min_units
                if max_units and total > max_units:
                    keep = False
                if keep:
                    frows.append(r)
                    if ratio:
                        frows.append(ratio)
                        i += 2
                    else:
                        i += 1
                else:
                    if ratio:
                        i += 2
                    else:
                        i += 1
            else:
                i += 1
        if frows:
            new_oh = sum(x["oh"] for x in frows if x.get("is_label_row"))
            new_wip = sum(x["wip"] for x in frows if x.get("is_label_row"))
            filtered.append({**block, "rows": frows, "total_oh": new_oh, "total_wip": new_wip})
    return filtered


def filter_categories(categories: list, min_units: int = 120, max_units: int = None) -> list:
    """Filter all categories' blocks. Recompute size_ranges after filtering."""
    result = []
    for cat in categories:
        fblocks = filter_blocks(cat["blocks"], min_units, max_units)
        if fblocks:
            # Recompute refs and OH/WIP per size range from filtered data
            sr_refs = defaultdict(list)
            sr_oh = defaultdict(int)
            sr_wip = defaultdict(int)

            for block in fblocks:
                for r in block["rows"]:
                    if r.get("is_label_row"):
                        ref = r.get("ref_num", "")
                        sr = r.get("size_range", "")
                        if sr and sr != "UNKNOWN" and ref:
                            if ref not in sr_refs[sr]:
                                sr_refs[sr].append(ref)
                            sr_oh[sr] += r["oh"]
                            sr_wip[sr] += r["wip"]

            # Build new size_ranges
            size_ranges = OrderedDict()
            all_sr_names = list(OrderedDict.fromkeys(
                list(sr_oh.keys()) + list(sr_refs.keys())
            ))
            for sr_name in all_sr_names:
                oh_val = sr_oh.get(sr_name, 0)
                wip_val = sr_wip.get(sr_name, 0)
                refs = sr_refs.get(sr_name, [])
                if oh_val > 0 or wip_val > 0 or refs:
                    size_ranges[sr_name] = {
                        "oh": oh_val, "wip": wip_val,
                        "total": oh_val + wip_val,
                        "refs": refs,
                    }

            result.append({
                **cat,
                "blocks": fblocks,
                "size_ranges": size_ranges,
            })
    return result


# --- Recap Data Builder --------------------------------------------------------

def get_recap_data(categories_by_sheet: dict) -> list:
    """Build recap data for the RECAP tab.

    Generates one row per size range per category (all 10 possible size ranges).
    """
    recap_sections = []
    for sheet_name, sheet_info in categories_by_sheet.items():
        brand = sheet_info["brand"]
        gen_cat = sheet_info.get("general_category", "")
        brand_label = f"{brand} {gen_cat}".strip() if gen_cat else brand

        section_rows = []
        total_oh, total_wip = 0, 0

        for cat_idx, cat in enumerate(sheet_info["categories"]):
            cat_name = cat["name"]
            # Unique ID per category entry so same-named categories stay separate
            cat_id = f"{sheet_name}_{cat_idx}"

            # One row per size range that exists for this category
            for sr_name, sr_data in cat["size_ranges"].items():
                sr_oh = sr_data["oh"]
                sr_wip = sr_data["wip"]
                ref_str = ", ".join(sr_data["refs"])
                section_rows.append({
                    "size_range": sr_name, "category": cat_name,
                    "cat_id": cat_id,
                    "ref_nums": ref_str, "oh": sr_oh, "wip": sr_wip,
                })
                total_oh += sr_oh
                total_wip += sr_wip

        recap_sections.append({
            "brand_label": brand_label, "brand": brand,
            "rows": section_rows,
            "total_oh": total_oh, "total_wip": total_wip,
            "total_ats": total_oh + total_wip,
        })
    return recap_sections
