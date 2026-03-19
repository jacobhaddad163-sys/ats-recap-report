"""
Excel output generator for ATS Recap Report.

Produces formatted .xlsx with:
  - Detail sheets (yellow category headers, grey sub-headers, summaries, images)
  - RECAP tab (brand summaries, size totals, grand total)
"""

import io
import logging
import re
from collections import OrderedDict
from datetime import date
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from .ats_parser import get_recap_data

logger = logging.getLogger(__name__)

_FORMULA_STARTERS = ('=', '+', '-', '@', '\t', '\r', '\n')

def _safe_cell_text(value: str) -> str:
    if not value:
        return value
    value = str(value).strip()
    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', value)
    if value and value[0] in _FORMULA_STARTERS:
        value = "'" + value
    return value

# ─── Style Constants ─────────────────────────────────────────────────────────
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
GREY_FILL = PatternFill("solid", fgColor="FFD3D3D3")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="C0E6F5")

BOLD_FONT = Font(name="Aptos Narrow", bold=True, size=11)
NORMAL_FONT = Font(name="Aptos Narrow", size=11)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
WRAP_CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

NUM_FMT = '#,##0'
ACCT_NUM_FMT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
PRICE_FMT = '#,##0.00'
TEXT_FMT = '@'


# ─── Detail Sheet ────────────────────────────────────────────────────────────

def _default_columns() -> dict:
    """Return the default (Nike/Jordan) column layout."""
    return {
        "style": 3, "color": 4, "oh": 12, "wip": 13, "avail": 14, "msrp": 15,
        "size_start": 5, "size_end": 11, "header_row": 0,
        "summary_label_col": 11,
    }


def _set_detail_col_widths(ws, cols: dict = None):
    if cols is None:
        cols = _default_columns()
    # Always set A and B for image/spacer area
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 6
    # Style and Color columns
    ws.column_dimensions[get_column_letter(cols["style"])].width = 22
    ws.column_dimensions[get_column_letter(cols["color"])].width = 28
    # Size columns
    for ci in range(cols["size_start"], cols["size_end"] + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 6
    # Summary label column (one before OH)
    ws.column_dimensions[get_column_letter(cols["summary_label_col"])].width = 10
    # OH, WIP, AVAIL, MSRP
    ws.column_dimensions[get_column_letter(cols["oh"])].width = 12
    ws.column_dimensions[get_column_letter(cols["wip"])].width = 10
    ws.column_dimensions[get_column_letter(cols["avail"])].width = 14
    ws.column_dimensions[get_column_letter(cols["msrp"])].width = 10


def _write_sheet_header(ws, report_date: date = None):
    if report_date is None:
        report_date = date.today()

    ws.merge_cells('A7:H7')
    cell = ws.cell(row=7, column=1, value="ATS RECAP")
    cell.font = Font(name="Aptos Narrow", bold=True, size=14)
    cell.alignment = LEFT_ALIGN

    ws.merge_cells('A8:H8')
    cell = ws.cell(row=8, column=1, value=report_date)
    cell.font = Font(name="Aptos Narrow", bold=True, size=11)
    cell.alignment = LEFT_ALIGN
    cell.number_format = 'M/D/YYYY'


def _write_category_summary(ws, row: int, label: str, oh: int, wip: int,
                             is_category_row: bool = False, category_name: str = "",
                             cols: dict = None):
    """Write a summary row (TODDLER or 4-7 line)."""
    if cols is None:
        cols = _default_columns()
    label_col = cols["summary_label_col"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    avail_col = cols["avail"]

    # Size range label column
    cell = ws.cell(row=row, column=label_col, value=label)
    cell.font = BOLD_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    # Labels like "4-7", "4-6X", "7-16", "8-20" need text format to prevent Excel date interpretation
    if any(ch.isdigit() for ch in label):
        cell.number_format = TEXT_FMT

    # OH — not bold, accounting format shows dash for zero
    cell = ws.cell(row=row, column=oh_col, value=oh)
    cell.font = NORMAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.number_format = ACCT_NUM_FMT
    cell.border = THIN_BORDER

    # WIP — not bold, accounting format shows dash for zero
    cell = ws.cell(row=row, column=wip_col, value=wip)
    cell.font = NORMAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.number_format = ACCT_NUM_FMT
    cell.border = THIN_BORDER

    # Total — not bold
    cell = ws.cell(row=row, column=avail_col, value=oh + wip)
    cell.font = NORMAL_FONT
    cell.alignment = CENTER_ALIGN
    cell.number_format = ACCT_NUM_FMT
    cell.border = THIN_BORDER

    # Category name in column A (yellow fill, bold)
    if is_category_row and category_name:
        cell = ws.cell(row=row, column=1, value=_safe_cell_text(category_name))
        cell.fill = YELLOW_FILL
        cell.font = BOLD_FONT
        cell.alignment = LEFT_ALIGN

    # Note: original raw file has NO borders on C:J of the 4-7 summary row


def _write_block_header(ws, row: int, cols: dict = None):
    """Write grey STYLE header row — matching original raw format exactly."""
    if cols is None:
        cols = _default_columns()
    style_col = cols["style"]
    color_col = cols["color"]
    size_start = cols["size_start"]
    size_end = cols["size_end"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    avail_col = cols["avail"]
    msrp_col = cols["msrp"]

    headers = {
        style_col: "STYLE", color_col: "COLOR", size_start: "SIZE SCALE",
        oh_col: "ON HAND", wip_col: "WIP", avail_col: "AVAILABILITY", msrp_col: "MSRP",
    }
    # Grey fill on cols from style to msrp (not 1-2 which are image area)
    for col in range(style_col, msrp_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = GREY_FILL
        c.font = NORMAL_FONT
        c.alignment = CENTER_ALIGN
    # Set header labels
    for col, label in headers.items():
        ws.cell(row=row, column=col, value=label)
    # Borders matching original pattern
    for col in [style_col, color_col, size_start, msrp_col]:
        ws.cell(row=row, column=col).border = THIN_BORDER
    for col in range(size_start + 1, size_end):
        ws.cell(row=row, column=col).border = Border(
            top=Side(style='thin'), bottom=Side(style='thin'))
    # Last size col: right+top+bottom (forms boundary before ON HAND)
    ws.cell(row=row, column=size_end).border = Border(
        right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in [oh_col, wip_col, avail_col]:
        ws.cell(row=row, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))


def _write_data_rows(ws, start_row: int, rows_data: list, cols: dict = None) -> int:
    """Write style data rows — matching original raw format exactly.

    Original border pattern:
      Label rows (has OH): thin on left, right, top — NO bottom
      Ratio rows (no OH):  thin on left, right, bottom — NO top
    Number format: General (matching raw file, no forced #,##0)
    """
    if cols is None:
        cols = _default_columns()
    style_col = cols["style"]
    color_col = cols["color"]
    size_start = cols["size_start"]
    size_end = cols["size_end"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    avail_col = cols["avail"]
    msrp_col = cols["msrp"]

    current_row = start_row
    for row_data in rows_data:
        is_label = row_data.get("is_label_row", True)

        ws.cell(row=current_row, column=style_col, value=row_data["style_num"]).font = NORMAL_FONT
        ws.cell(row=current_row, column=color_col, value=row_data["color"]).font = NORMAL_FONT

        for col_idx in range(size_start, size_end + 1):
            val = row_data["cells"].get(col_idx)
            if val is not None:
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = NORMAL_FONT
                cell.alignment = CENTER_ALIGN

        if is_label and row_data["oh"] > 0:
            cell = ws.cell(row=current_row, column=oh_col, value=row_data["oh"])
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN

        if is_label and row_data["wip"] > 0:
            cell = ws.cell(row=current_row, column=wip_col, value=row_data["wip"])
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN

        avail = row_data.get("availability", "")
        if avail:
            ws.cell(row=current_row, column=avail_col, value=avail).font = NORMAL_FONT

        msrp = row_data.get("msrp", 0)
        if msrp > 0:
            cell = ws.cell(row=current_row, column=msrp_col, value=msrp)
            cell.font = NORMAL_FONT
            cell.alignment = CENTER_ALIGN
            cell.number_format = PRICE_FMT

        # Borders — match original raw file pattern per column
        if is_label:
            for col in [style_col, color_col]:
                ws.cell(row=current_row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
            ws.cell(row=current_row, column=size_start).border = Border(
                left=Side(style='thin'), top=Side(style='thin'))
            for col in range(size_start + 1, size_end):
                ws.cell(row=current_row, column=col).border = Border(top=Side(style='thin'))
            # Last size col: right border only (boundary before ON HAND)
            ws.cell(row=current_row, column=size_end).border = Border(
                right=Side(style='thin'))
            for col in [oh_col, wip_col, avail_col, msrp_col]:
                ws.cell(row=current_row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
        else:
            for col in [style_col, color_col]:
                ws.cell(row=current_row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=current_row, column=size_start).border = Border(
                left=Side(style='thin'), bottom=Side(style='thin'))
            for col in range(size_start + 1, size_end):
                ws.cell(row=current_row, column=col).border = Border(bottom=Side(style='thin'))
            # Last size col: right+bottom border
            ws.cell(row=current_row, column=size_end).border = Border(
                right=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=current_row, column=oh_col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'))
            for col in [wip_col, avail_col]:
                ws.cell(row=current_row, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
            ws.cell(row=current_row, column=msrp_col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))

        current_row += 1
    return current_row


def _write_total_row(ws, row: int, total_oh: int, total_wip: int, cols: dict = None):
    """Write TOTAL row — matching original raw format exactly.

    Original: grey fill on style-to-msrp cols only, TOTAL text bold, values NOT bold,
    General number format, specific border pattern.
    """
    if cols is None:
        cols = _default_columns()
    style_col = cols["style"]
    color_col = cols["color"]
    size_start = cols["size_start"]
    size_end = cols["size_end"]
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    avail_col = cols["avail"]
    msrp_col = cols["msrp"]

    # Grey fill on cols style-to-msrp (not 1-2 image area)
    for col in range(style_col, msrp_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = GREY_FILL
        c.font = NORMAL_FONT

    # TOTAL text — bold
    cell = ws.cell(row=row, column=style_col, value="TOTAL :")
    cell.font = BOLD_FONT

    # OH — not bold, General format
    cell = ws.cell(row=row, column=oh_col, value=total_oh)
    cell.alignment = CENTER_ALIGN

    # WIP
    if total_wip > 0:
        cell = ws.cell(row=row, column=wip_col, value=total_wip)
        cell.alignment = CENTER_ALIGN

    # Borders matching original pattern
    for col in [style_col, color_col, msrp_col]:
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=size_start).border = Border(
        left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(size_start + 1, size_end + 1):
        ws.cell(row=row, column=col).border = Border(
            top=Side(style='thin'), bottom=Side(style='thin'))
    for col in [oh_col, wip_col, avail_col]:
        ws.cell(row=row, column=col).border = Border(
            left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


def _add_product_image(ws, row: int, img_bytes: bytes):
    """Add product image at the given row — resize to match original placement.

    Original images span ~10 rows in column A (twoCellAnchor, ~200x200px).
    openpyxl ignores width/height setters when serializing — it uses the
    actual image pixel data. So we must resize the image bytes with PIL first.
    """
    try:
        from PIL import Image as PILImage
        pil_img = PILImage.open(io.BytesIO(img_bytes))
        w, h = pil_img.size
        max_w, max_h = 200, 200
        if w > max_w or h > max_h:
            scale = min(max_w / w, max_h / h)
            new_w, new_h = int(w * scale), int(h * scale)
            pil_img = pil_img.resize((new_w, new_h), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil_img.save(buf, format='PNG')
        buf.seek(0)
        img = XlImage(buf)
        ws.add_image(img, f'A{row}')
    except Exception:
        # Fallback: use original size if PIL unavailable
        try:
            img = XlImage(io.BytesIO(img_bytes))
            ws.add_image(img, f'A{row}')
        except Exception:
            pass


def write_detail_sheet(ws, categories: list, report_date: date = None,
                       cols: dict = None):
    """
    Write a complete detail sheet.
    categories: list of category dicts from parser (with toddler_oh, boys47_oh, blocks, etc.)
    cols: detected column layout dict from parser (or None for defaults)
    """
    if cols is None:
        cols = _default_columns()
    oh_col = cols["oh"]
    wip_col = cols["wip"]
    avail_col = cols["avail"]
    label_col = cols["summary_label_col"]

    _set_detail_col_widths(ws, cols)
    _write_sheet_header(ws, report_date)
    current_row = 10

    for cat in categories:
        cat_name = cat["name"]
        size_ranges = cat.get("size_ranges", OrderedDict())
        blocks = cat["blocks"]

        if not blocks:
            continue

        # OH/WIP/TOTAL column sub-headers (yellow) — aligned above data columns
        for col, label in [(oh_col, "OH"), (wip_col, "WIP"), (avail_col, "TOTAL")]:
            cell = ws.cell(row=current_row, column=col, value=label)
            cell.fill = YELLOW_FILL
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        # Summary label col also gets yellow fill and border (no label)
        ws.cell(row=current_row, column=label_col).fill = YELLOW_FILL
        ws.cell(row=current_row, column=label_col).border = THIN_BORDER
        current_row += 1

        # Write one summary row per size range; category name goes on the LAST row
        sr_items = list(size_ranges.items())
        for idx, (sr_name, sr_data) in enumerate(sr_items):
            is_last = (idx == len(sr_items) - 1)
            label = sr_name  # Full name: "TODDLER BOY", "4-7 BOY", "INFANT GIRL", etc.
            _write_category_summary(
                ws, current_row, label, sr_data["oh"], sr_data["wip"],
                is_category_row=is_last, category_name=cat_name if is_last else "",
                cols=cols,
            )
            current_row += 1

        # Write all style blocks for this category
        for block in blocks:
            block_header_row = current_row
            _write_block_header(ws, current_row, cols=cols)
            # Place image at the STYLE header row in col A — matching original
            if block.get("product_image"):
                _add_product_image(ws, block_header_row, block["product_image"])
            current_row += 1
            current_row = _write_data_rows(ws, current_row, block["rows"], cols=cols)
            _write_total_row(ws, current_row, block["total_oh"], block["total_wip"], cols=cols)
            current_row += 1
            current_row += 8  # spacing after TOTAL (image area)

        current_row += 2


# ─── RECAP Sheet ─────────────────────────────────────────────────────────────

def _set_recap_col_widths(ws):
    widths = {'A': 20.5, 'B': 15.3, 'C': 26.0, 'D': 35.2, 'E': 9.9, 'F': 9.9, 'G': 9.9}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_recap_sheet(ws, recap_sections: list, title: str = ""):
    _set_recap_col_widths(ws)

    # Collect merge ranges — apply them at the very end so all cell styling
    # happens on real Cell objects (not MergedCell proxies).
    pending_merges = []

    # Row 1: Title — style ALL cells in the range before merging
    for col in range(1, 8):
        c = ws.cell(row=1, column=col)
        c.fill = YELLOW_FILL
        c.font = BOLD_FONT
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER
    ws.cell(row=1, column=1, value=_safe_cell_text(title.upper() if title else "ATS RECAP"))
    pending_merges.append('A1:G1')

    # Row 2: Headers
    headers = ["BRAND", "SIZE RANGE", "CATEGORY", "REF #", "OH", "WIP", "TOTAL ATS"]
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.fill = YELLOW_FILL
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    current_row = 3
    all_data_rows = []
    brand_total_rows = []

    for section in recap_sections:
        brand_label = section["brand_label"]
        rows = section["rows"]
        if not rows:
            continue

        section_start_row = current_row

        # Group rows by category for merging.
        # Use cat_id (unique per category entry) to keep same-named categories
        # separate — spec says duplicate category names are SEPARATE entries.
        cat_groups = []
        current_cat_id = None
        for row_data in rows:
            cat_id = row_data.get("cat_id")
            if cat_id != current_cat_id:
                cat_groups.append([row_data])
                current_cat_id = cat_id
            else:
                cat_groups[-1].append(row_data)

        for cat_group in cat_groups:
            cat_start_row = current_row

            for row_data in cat_group:
                # Apply white fill and borders to all cells in this data row
                for col in range(1, 8):
                    c = ws.cell(row=current_row, column=col)
                    c.fill = WHITE_FILL
                    c.border = THIN_BORDER

                # B: SIZE RANGE
                cell_b = ws.cell(row=current_row, column=2, value=row_data["size_range"])
                cell_b.font = NORMAL_FONT
                cell_b.alignment = CENTER_ALIGN

                # C: Category name (set on every row — merge later collapses to anchor)
                ws.cell(row=current_row, column=3).font = NORMAL_FONT
                ws.cell(row=current_row, column=3).alignment = WRAP_CENTER_ALIGN

                # D: REF # — text format (@) with wrap_text
                cell_d = ws.cell(row=current_row, column=4,
                                 value=_safe_cell_text(row_data["ref_nums"]))
                cell_d.font = NORMAL_FONT
                cell_d.alignment = WRAP_CENTER_ALIGN
                cell_d.number_format = TEXT_FMT

                # E: OH
                cell = ws.cell(row=current_row, column=5, value=row_data["oh"])
                cell.font = NORMAL_FONT
                cell.alignment = CENTER_ALIGN
                cell.number_format = NUM_FMT

                # F: WIP
                cell = ws.cell(row=current_row, column=6, value=row_data["wip"])
                cell.font = NORMAL_FONT
                cell.alignment = CENTER_ALIGN
                cell.number_format = NUM_FMT

                # G: TOTAL ATS (formula)
                cell = ws.cell(row=current_row, column=7, value=f'=E{current_row}+F{current_row}')
                cell.font = NORMAL_FONT
                cell.alignment = CENTER_ALIGN
                cell.number_format = NUM_FMT

                all_data_rows.append(current_row)
                current_row += 1

            # Category name value (on anchor cell)
            ws.cell(row=cat_start_row, column=3,
                    value=_safe_cell_text(cat_group[0]["category"]))

            # Queue category merge
            if len(cat_group) > 1:
                pending_merges.append(f'C{cat_start_row}:C{current_row - 1}')

        # Brand name value (on anchor cell)
        if current_row > section_start_row:
            cell = ws.cell(row=section_start_row, column=1, value=_safe_cell_text(brand_label))
            cell.font = BOLD_FONT
            cell.alignment = WRAP_CENTER_ALIGN
            # Queue brand merge
            if current_row - section_start_row > 1:
                pending_merges.append(f'A{section_start_row}:A{current_row - 1}')

        # Brand total row (light blue) — style ALL cells A:G before merging
        for col in range(1, 8):
            c = ws.cell(row=current_row, column=col)
            c.fill = LIGHT_BLUE_FILL
            c.border = THIN_BORDER

        ws.cell(row=current_row, column=1,
                value=_safe_cell_text(f"{brand_label} TOTAL:"))
        ws.cell(row=current_row, column=1).font = BOLD_FONT
        ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN

        for col_idx, formula in [(5, f'=SUM(E{section_start_row}:E{current_row - 1})'),
                                  (6, f'=SUM(F{section_start_row}:F{current_row - 1})'),
                                  (7, f'=E{current_row}+F{current_row}')]:
            cell = ws.cell(row=current_row, column=col_idx, value=formula)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            cell.number_format = NUM_FMT

        pending_merges.append(f'A{current_row}:D{current_row}')
        brand_total_rows.append(current_row)
        current_row += 1

    # Size range totals (TODDLER TOTAL, 4-7 TOTAL, etc.)
    unique_srs = []
    for section in recap_sections:
        for row_data in section["rows"]:
            if row_data["size_range"] not in unique_srs:
                unique_srs.append(row_data["size_range"])

    sr_sort_order = [
        "NB GIRL", "INFANT GIRL", "TODDLER GIRL", "4-6X GIRL", "7-16 GIRL",
        "NB BOY", "INFANT BOY", "TODDLER BOY", "4-7 BOY", "8-20 BOY",
    ]
    unique_srs.sort(key=lambda x: sr_sort_order.index(x) if x in sr_sort_order else 99)

    first_data = min(all_data_rows) if all_data_rows else 3
    last_data = max(all_data_rows) if all_data_rows else 3

    for sr_name in unique_srs:
        # Style ALL cells A:G before merging
        for col in range(1, 8):
            c = ws.cell(row=current_row, column=col)
            c.fill = LIGHT_BLUE_FILL
            c.border = THIN_BORDER

        ws.cell(row=current_row, column=1, value=_safe_cell_text(f"{sr_name} TOTAL"))
        ws.cell(row=current_row, column=1).font = BOLD_FONT
        ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN

        safe_sr = re.sub(r'["\';=+@]', '', sr_name)

        for col_idx, formula in [
            (5, f'=SUMIF($B${first_data}:$B${last_data},"{safe_sr}",E${first_data}:E${last_data})'),
            (6, f'=SUMIF($B${first_data}:$B${last_data},"{safe_sr}",F${first_data}:F${last_data})'),
            (7, f'=E{current_row}+F{current_row}'),
        ]:
            cell = ws.cell(row=current_row, column=col_idx, value=formula)
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            cell.number_format = NUM_FMT

        pending_merges.append(f'A{current_row}:D{current_row}')
        current_row += 1

    # Grand Total (yellow) — style ALL cells before merging
    for col in range(1, 8):
        c = ws.cell(row=current_row, column=col)
        c.fill = YELLOW_FILL
        c.border = THIN_BORDER

    ws.cell(row=current_row, column=1, value="GRAND TOTAL:")
    ws.cell(row=current_row, column=1).font = BOLD_FONT
    ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN

    oh_formula = "+".join(f"E{r}" for r in brand_total_rows) if brand_total_rows else "0"
    wip_formula = "+".join(f"F{r}" for r in brand_total_rows) if brand_total_rows else "0"

    for col_idx, formula in [(5, f'={oh_formula}'), (6, f'={wip_formula}'),
                              (7, f'=E{current_row}+F{current_row}')]:
        cell = ws.cell(row=current_row, column=col_idx, value=formula)
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.number_format = NUM_FMT

    pending_merges.append(f'A{current_row}:D{current_row}')

    # Auto-fit row heights for REF# column (D) — openpyxl doesn't auto-fit,
    # so estimate based on content length vs column width.
    col_d_width = 35.2  # character units
    char_per_line = max(int(col_d_width * 0.85), 1)  # rough chars that fit per line
    default_row_ht = 15  # default Excel row height in points
    line_ht = 15  # points per line of text

    for row_idx in all_data_rows:
        cell_d = ws.cell(row=row_idx, column=4)
        val = str(cell_d.value or '')
        if not val:
            continue
        # Estimate lines needed
        lines = max(1, -(-len(val) // char_per_line))  # ceiling division
        needed_ht = lines * line_ht
        if needed_ht > default_row_ht:
            ws.row_dimensions[row_idx].height = needed_ht

    # NOW apply all merges — after all cell styling is complete
    for merge_range in pending_merges:
        ws.merge_cells(merge_range)


# ─── Main Generator ─────────────────────────────────────────────────────────

def generate_ats_report(categories_by_sheet: Dict[str, dict],
                        title: str = "", report_date: date = None,
                        logo_image: bytes = None) -> bytes:
    """
    Generate the complete ATS report Excel file.

    categories_by_sheet: {
        "NIKE LONG BOTTOMS 2-7": {
            "brand": "NIKE",
            "general_category": "LONG BOTTOMS",
            "categories": OrderedDict of {cat_name: {"blocks": [...], "size_ranges": {...}}},
        },
        ...
    }
    logo_image: bytes of the Haddad Brands logo (placed at A1 on each detail sheet)
    """
    wb = Workbook()
    ws_recap = wb.active
    ws_recap.title = "RECAP SHEET"

    recap_sections = get_recap_data(categories_by_sheet)
    write_recap_sheet(ws_recap, recap_sections, title=title)

    for sheet_name, sheet_info in categories_by_sheet.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        sheet_cols = sheet_info.get("columns")
        write_detail_sheet(ws, sheet_info["categories"], report_date=report_date,
                           cols=sheet_cols)
        # Add Haddad Brands logo at A1 (same position as original)
        if logo_image:
            try:
                img = XlImage(io.BytesIO(logo_image))
                ws.add_image(img, 'A1')
            except Exception:
                pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
